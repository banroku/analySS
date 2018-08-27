# -*- coding: utf-8 -*-
"""
Created on Thu Jul 12 16:41:18 2018

@author: ppc213937
"""

import openpyxl as pyxl
import collections

file = './test/formula.xlsx'
wb_formura = pyxl.load_workbook(file)
ws_formura = wb_formura.worksheets[0]


''' get final row '''
row_final = 1
for i in range (1, 10000):
    if ws_formura.cell(i, 1).value == None:
        row_final = i - 1
        break

''' get list of addtive '''
additive_list = []
for i in range(6, 21, 2): 
    for j in range(1, row_final):
        additive_list.append(ws_formura.cell(j, i).value)

''' get unique list of addtive '''
additive_uniquelist = list(set(additive_list))

''' to do. create list of uniquelist vs list '''
wb_new = pyxl.Workbook()
for i in range(1, len(additive_uniquelist)):
    wb_new.worksheets[0].cell(i,1).value = additive_uniquelist[i-1]

